'''
Português
1 - Entrar na planilha, confirmar a data e extrair o serviço
2 - Faço login pra ter acesso a página: http://backofficevendaembarcada.rodosoft.com.br/Pages/Operacoes, uso o serviço para pesquisar os dados
    2.1 - os dados necessarios são serviço, data e hora, inicio da viagem e fim da viagem
3 - Inserir os dados em uma nova planilha
4 - Repetir até terminar os serviços do dia de hoje
'''

'''
English
1 - Enter the spreadsheet, search the services of the right day
2 - Login on the page http://backofficevendaembarcada.rodosoft.com.br/Pages/Operacoes, and use the service to catch necessary data
    2.1 - necessary data = service, dayHour, travelStart, travelEnd
3 - After the search, insert the data in another spreadsheet
4 - Repeat in all services in that day
'''

import openpyxl
import openpyxl.workbook
import os
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from datetime import datetime, timedelta

searchDay = lambda: (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")

def login():
    boxKey = driver.find_element(By.XPATH, "//input[@id='chave']")
    boxLogin = driver.find_element(By.XPATH, "//input[@id='login']")
    boxPassword = driver.find_element(By.XPATH, "//input[@id='password']")
    btnLogin = driver.find_element(By.XPATH, "//input[@class='btn btn-block btn-lg btn-danger']")
    boxKey.send_keys("insira a chave aqui")
    boxLogin.send_keys("insira o login")
    boxPassword.send_keys("insira a senha")
    sleep(0.5)
    btnLogin.click()
    driver.get('http://backofficevendaembarcada.rodosoft.com.br/Pages/Operacoes')
    sleep(1)

def search():
    boxService = driver.find_element(By.XPATH, "//input[@id='MainContent_txtServico']")
    btnSearch = driver.find_element(By.XPATH, "//input[@id='MainContent_Button1']")
    boxDate = driver.find_element(By.XPATH, "//input[@id='MainContent_btnenableddate']")
    sleep(0.5)
    boxService.clear()
    boxService.send_keys(CORRIDA_ID)
    boxDate.clear()
    boxDate.send_keys(DATA_SERVIÇO)
    sleep(0.2)
    btnSearch.click()
    sleep(0.5)

def dataMining(service, dayHour, travelStart, travelEnd):
    td_elements = driver.find_elements(By.XPATH, "//table[@class='table margin table-striped table-hover sources-table']//td")
    sleep(0.5)
    service = td_elements[0].text.strip()
    dayHour = td_elements[5].text.strip()
    travelStart = td_elements[6].text.strip()
    travelEnd = td_elements[8].text.strip()
    sleep(0.5)
    return service, dayHour, travelStart, travelEnd

def save(filePathSave):
    resultSheet = openpyxl.load_workbook(filePathSave)
    pageSheet = resultSheet['Sheet1']
    pageSheet.append([service, dayHour, travelStart, travelEnd])
    resultSheet.save(filePathSave)

def renameAndMoveFile():
    filePathSave = os.path.abspath('planilhaDiariaMonitriip.xlsx')
    saveDay = datetime.now() - timedelta(days=1) 
    newFileName = f'planilhaDiariaMonitriip{saveDay.strftime("%d-%m-%Y")}.xlsx'
    os.rename(filePathSave, newFileName)
    shutil.move(newFileName, os.path.join('planilhasGeradas', newFileName))

def generateNewFile():
    newFile = openpyxl.Workbook()
    sheet = newFile.active
    sheet.title = 'Sheet1'
    headers = ['SERVIÇO', 'HORA_PLANEJADA', 'HORA_INICIADA', 'HORA_FINALIZADA']
    sheet.append(headers)
    newFile.save('planilhaDiariaMonitriip.xlsx')


# 1 - Entrar na planilha

filePathServices = os.path.abspath('planilhaServicos2024.xlsx')
sheetServices = openpyxl.load_workbook(filePathServices)
pageServices = sheetServices['Sheet1']

# 2 - Fazer login pra ter acesso a página: http://backofficevendaembarcada.rodosoft.com.br/Pages/Operacoes, onde usar o serviço para pesquisar

driver = webdriver.Chrome()
driver.get('http://backofficevendaembarcada.rodosoft.com.br/Pages/Login')
login()

# 3 - Verifica a data na planilha e coloca o serviço até acabarem os serviços na data de hoje 
for line in pageServices.iter_rows(min_row=2, values_only=True):    
    DATA_SERVIÇO, CORRIDA_ID = line
    if DATA_SERVIÇO == searchDay():

        search()
        service, dayHour, travelStart, travelEnd = "", "", "", "" 
        service, dayHour, travelStart, travelEnd = dataMining(service, dayHour, travelStart, travelEnd)

        filePathSave = os.path.abspath('planilhaDiariaMonitriip.xlsx')
        save(filePathSave)

renameAndMoveFile()

generateNewFile()
